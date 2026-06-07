#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Converte as planilhas do sistema anterior (Clientes.xlsx / Dispositivos.xlsx)
para JSON limpo em UTF-8, consumido por import-dados-anteriores.ts.

Uso:
    pip install openpyxl
    python converter-xlsx.py

As planilhas sao lidas de ../../../DadosDoSistemaAnterior por padrao
(raiz do repositorio). Os JSONs sao gravados em ./data.
"""
import json
import os
import sys

import openpyxl

AQUI = os.path.dirname(os.path.abspath(__file__))
# backend/scripts/importar -> sobe 3 niveis ate a raiz do repo
RAIZ = os.path.abspath(os.path.join(AQUI, "..", "..", ".."))
ORIGEM = os.environ.get("ORIGEM_DADOS", os.path.join(RAIZ, "DadosDoSistemaAnterior"))
DESTINO = os.path.join(AQUI, "data")


def s(v):
    """Normaliza celula -> string trimada ou None."""
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def ler_clientes(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb.active
    # Colunas: Nome, Email, Desativado, Telefone
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nome = s(r[0])
        if not nome:
            continue
        out.append({
            "nome": nome,
            "email": s(r[1]),
            "desativado": s(r[2]),  # "Não" / "Sim"
            "telefone": s(r[3]),
        })
    return out


# indice -> chave do JSON para Dispositivos.xlsx
COLS_DISP = {
    0: "nome", 1: "identificador", 2: "modelo", 3: "grupo", 4: "categoria",
    5: "contato", 6: "desativado", 7: "telefone", 8: "placa", 9: "marca",
    10: "modeloVeiculo", 11: "combustivel", 12: "consumo", 13: "iccid",
    14: "operadora", 15: "cor", 16: "ano", 17: "renavam", 18: "chassi",
    19: "localInstalacao", 20: "instalador", 21: "senha",
    # 22 = Ultima Atualizacao (ignorado)
}


def ler_dispositivos(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True)
    ws = wb.active
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        ident = s(r[1]) if len(r) > 1 else None
        if not ident:
            continue
        item = {}
        for idx, chave in COLS_DISP.items():
            item[chave] = s(r[idx]) if idx < len(r) else None
        out.append(item)
    return out


def main():
    os.makedirs(DESTINO, exist_ok=True)
    f_cli = os.path.join(ORIGEM, "Clientes.xlsx")
    f_disp = os.path.join(ORIGEM, "Dispositivos.xlsx")

    if not os.path.exists(f_cli) or not os.path.exists(f_disp):
        print(f"ERRO: planilhas nao encontradas em {ORIGEM}", file=sys.stderr)
        print("Defina ORIGEM_DADOS apontando para a pasta das planilhas.", file=sys.stderr)
        sys.exit(1)

    clientes = ler_clientes(f_cli)
    dispositivos = ler_dispositivos(f_disp)

    with open(os.path.join(DESTINO, "clientes.json"), "w", encoding="utf-8") as fp:
        json.dump(clientes, fp, ensure_ascii=False, indent=2)
    with open(os.path.join(DESTINO, "dispositivos.json"), "w", encoding="utf-8") as fp:
        json.dump(dispositivos, fp, ensure_ascii=False, indent=2)

    print(f"OK: {len(clientes)} clientes -> data/clientes.json")
    print(f"OK: {len(dispositivos)} dispositivos -> data/dispositivos.json")


if __name__ == "__main__":
    main()
